# Code by Elisabeth Holm
# Analyzes chromatophore data to connect
# temporal and spatial chromatophore data
# Python 3.10.4
# March 2023 - Present

import math
from argparse import ArgumentDefaultsHelpFormatter, ArgumentParser  # argparse 1.4.0
from gooey import Gooey, GooeyParser  # Gooey 1.0.8.1
import numpy as np  # numpy 1.22.4
import pandas as pd  # pandas 1.5.3
import matplotlib.pyplot as plt  # matplotlib 3.5.2
from matplotlib import cm, colors
import os  # for finding context image
import cv2  # for drawing pathways image -- can prob replace w smaller module later
import openpyxl  # openpyxl 3.1.1

filepath = "New_Vids/Sept_2022_Tests/TS-20220801155100769_367_12_383_456.xlsx"  # clear chrom pathway, ROI near electrode
chrom_IDs = []  # list of chrom IDs in order of how they appear on spreadsheet, populate in main()
args = []
window_size = 30
n_df = None  # df from neighborhood sheet (populated in load())
a_df = None  # df from activations sheet  (populated in load())
c_df = None  # df from centroids sheet  (populated in load())


@Gooey  # The GUI Decorator goes here
def parse_args():
    global args

    ap = GooeyParser(formatter_class=ArgumentDefaultsHelpFormatter,
                     conflict_handler='resolve')

    ap.add_argument("-s", "--spreadsheet", default=filepath, widget="FileChooser",
                    help="File path of the spreadsheet (.xlsx file generated from pathways.py)")
    ap.add_argument("-w", "--window_size", default=window_size, type=int,
                    help="Size of the time window to divide frames by")
    ap.add_argument("-h", "--display_hist", default="off", choices=["off", "on"],
                    help="Display histogram of chromatophore activations over time")
    ap.add_argument("-a", "--arrows_img", default="auto_detect", widget="FileChooser",
                    help="File path of the context image (to annotate with activation pathway arrows)")
    ap.add_argument("-c", "--arrow_color", default="rainbow spectrum",
                    choices=["uniform (green/blue)", "rainbow spectrum",
                             "cyan/magenta spectrum", "blue/green spectrum"],
                    help="Color of arrows that draw activation pathway")
    ap.add_argument("-r", "--activations_of_interest", default="", type=str,
                    help="start and end for activation numbers of interest, each number separated "
                         "by a space (e.g. 3 8). Useful to look at fewer activations if "
                         "many activations overlap in the pathways image. "
                         "Leave blank to draw all activations on the pathways image.")

    args = vars(ap.parse_args())


# load the relevant sheets from the file
def load(filename):
    # neighborhood spreadsheet
    n_df = pd.read_excel(f"{filename}", sheet_name="Neighbors")
    a_df = pd.read_excel(f"{filename}", sheet_name="Activations")
    c_df = pd.read_excel(f"{filename}", sheet_name="Centroids")

    return n_df, a_df, c_df


# set global nums based on args
def set_global_nums():
    global filepath
    global n_df
    global a_df
    global c_df
    global chrom_IDs
    global window_size

    filepath = args["spreadsheet"]
    window_size = args["window_size"]
    n_df, a_df, c_df = load(filepath)  # load the relevant sheets from the file
    chrom_IDs = a_df.loc[:,"Chromatophore ID"].values.tolist()  # get list of chrom IDs


# sort all activations chronologically
# return 2D list of activations in chronological order with items [chrom_ID, activation_frame]
# e.g [["C12", 23], ["C45", 28], ...]
# also return activations sorted by chromatophore
# e.g. {"C12": [23, 48], "C45":[28]}
def get_sorted_activations():
    chronological_activations = []
    activations_by_chrom = {}

    # iterate through each chromatophore, populate sorted_activations list
    for i in range(len(chrom_IDs)):
        row = a_df.iloc[i]
        chrom_ID = row[0]
        chrom_activations = []
        # iterate through all activation frames
        for a_i in range(1, len(row), 3):
            a_frame = row[a_i]
            # if this is an activation (not nan)
            if not row.isnull()[a_i]:
                chronological_activations.append([chrom_ID, a_frame])
                chrom_activations.append(a_frame)

        activations_by_chrom[chrom_ID] = chrom_activations

    # sort chronologically
    chronological_activations = sorted(chronological_activations, key=lambda l:l[1])

    return chronological_activations, activations_by_chrom


# sort activation frame num into windows
# return a dataframe with column names as chromatophore and rows as windows of frame nums
def make_act_histogram(activations):
    n_frames = len(pd.read_excel(f"{filepath}", sheet_name="Areas"))  # number of frames in vid
    n_windows = int(n_frames/window_size) + 1
    windows = np.linspace(0, n_frames, n_windows)  # number will be start of window (0-29, 30-59, etc)
    data = {"window start": windows}  # initialize data with windows as first column

    # iterate through chronological activations
    for chrom_ID, act_list in activations.items():
        chrom_bin_hist_data = [0]*len(windows)  # initialize a list of 0's

        # iterate through each activation, switch bit of whatever window it happens in to 1
        for act in act_list:
            window_i = int(act // window_size)  # index of window that activation falls under
            chrom_bin_hist_data[window_i] = 1  # indicate this chrom activates in this window

        data[chrom_ID] = chrom_bin_hist_data  # store column of data in data

    hist_a_df = pd.DataFrame(data)

    if args["display_hist"] == "on":
        plot_act_histogram(hist_a_df, window_size)

    return hist_a_df


# plot histogram of activations
def plot_act_histogram(hist_a_df, window_length):
    num_windows = len(hist_a_df)
    windows = []
    num_window_activations = [0] * num_windows

    # iterate through all rows of hist_a_df
    for i in range(num_windows):
        row = hist_a_df.iloc[i]  # data for this window
        windows.append(str(int(row[0])) + "-" + str(int(row[0]) + window_length - 1))  # bar label
        win_data = row[1:]
        num_window_activations[i] = win_data.sum()  # number of chroms that activated in this window

    # plot graph
    plt.bar(windows, num_window_activations)
    plt.title('Number of chromatophore activations per frame window', fontsize=14)
    plt.xlabel('Range of Frames', fontsize=14)
    plt.ylabel('Number of Activations', fontsize=14)
    plt.grid(True)
    plt.show()


# P(A_i AND A_j)
# = # windows where A_i activates AND A_j activates / # windows
def get_p_A_i_and_A_j(df, ID_i, ID_j):
    rows_Ai_and_Aj = df.loc[(df[ID_i] == 1) & (df[ID_j] == 1)]
    count_Ai_and_Aj = len(rows_Ai_and_Aj)
    count_windows = len(df)

    p_A_i_and_A_j = count_Ai_and_Aj / count_windows
    return p_A_i_and_A_j


# P(A_i)
# = # windows where A_i activates / # windows
def get_p_A_i(df, ID_i):
    rows_Ai = df.loc[df[ID_i] == 1]
    count_Ai = len(rows_Ai)
    count_windows = len(df)

    p_A_i = count_Ai / count_windows
    return p_A_i


# return bool whether a given pair is dependent from one another
# in temporal activation data
# (if True, their activations happen somewhat at the same time usually)
def pair_time_dependent(hist_a_df, ID_i, ID_j):
    p_A_i_and_A_j = get_p_A_i_and_A_j(hist_a_df, ID_i, ID_j)  # P(A_i AND A_j)
    p_A_i = get_p_A_i(hist_a_df, ID_i)  # P(A_i)
    p_A_j = get_p_A_i(hist_a_df, ID_j)  # P(A_j)

    # if P(A_i AND A_j) = P(A_i)P(A_j), pair is independent
    difference = abs(p_A_i_and_A_j - p_A_i*p_A_j)
    # also don't count pair as dep if one of them never activated
    if difference < 0.01 or p_A_i*p_A_j == 0:
        dep = False  # they are independent
    else:
        dep = True  # they are dependent

    # print calculations for ID_i and ID_j being independent based on activation time(s)
    # print("Chrom_i:", ID_i, ", Chrom_j:", ID_j)
    # print("P(A_i):", p_A_i)
    # print("P(A_j):", p_A_j)
    # print("P(A_i AND A_j):", p_A_i_and_A_j)
    # print("P(A_i)*P(A_j):", p_A_i*p_A_j)
    # print("Difference:", difference)
    # print("Dependent:", dep)
    # print()

    return dep


# find all pairs that are time dependent on each other
# return dict with key chrom_ID : value {chrom IDs that this one is dependent with}
def get_time_dep_pairs(hist_a_df):
    time_dep = {}
    num_time_dep_pairs = 0

    # initialize time_dep dict with empty sets
    for chrom_ID in chrom_IDs:
        time_dep[chrom_ID] = set()

    # print("If A_i and A_j are independent, we expect P(A_i AND A_j) = P(A_i) * P(A_j)")
    # print()

    # check every pair
    for i in range(len(chrom_IDs)):
        ID_i = chrom_IDs[i]
        for j in (range(i + 1, len(chrom_IDs))):
            ID_j = chrom_IDs[j]

            # if pair is time dependent on each other
            if pair_time_dependent(hist_a_df, ID_i, ID_j):
                # add both chroms as dependent of each other
                time_dep[ID_i].add(ID_j)
                time_dep[ID_j].add(ID_i)
                num_time_dep_pairs += 1  # keep track of how many time_dep_pairs for P(T) later

    return time_dep, num_time_dep_pairs


# get neighborhood as a dict with key chrom_ID: value {neighbors}
def get_neighborhood():
    neighborhood = {}

    # initialize neighborhood dict with empty sets
    for chrom_ID in chrom_IDs:
        neighborhood[chrom_ID] = set()

    # iterate through all rows in Neighbors spreadsheet
    for i in range(len(chrom_IDs)):
        row = n_df.iloc[i]
        chrom_ID = row[0]
        # iterate through all neighbors
        for n_i in range(1, len(row)):
            # if the value isn't NaN
            if not row.isnull()[n_i]:
                neighborhood[chrom_ID].add(row[n_i])

    return neighborhood


# P(T AND N)
# # dependent neighbor pairs / # pairs
def get_p_T_and_N(time_dep, neighbors, num_pairs):
    num_T_and_N = 0

    # check every pair
    for i in range(len(chrom_IDs)):
        ID_i = chrom_IDs[i]
        for j in (range(i + 1, len(chrom_IDs))):
            ID_j = chrom_IDs[j]

            # if ID_j is both time-dependent and a neighbor of ID_i
            if ID_j in time_dep[ID_i] and ID_j in neighbors[ID_i]:
                num_T_and_N += 1

    p_T_and_N = num_T_and_N / num_pairs
    print("P(T AND N) = # pairs that are neighbors and time-dependent / # pairs")
    print("=", num_T_and_N, "/", num_pairs, "=", p_T_and_N)
    print()
    return p_T_and_N


# P(T)
# # time-dependent pairs / # pairs
def get_p_T(num_time_dep_pairs, num_pairs):
    p_T = num_time_dep_pairs / num_pairs
    print("P(T) = # time-dependent pairs / # pairs")
    print("=", num_time_dep_pairs, "/", num_pairs, "=", p_T)
    print()
    return p_T


# P(N)
# # neighboring pairs / # pairs
def get_p_N(neighborhood, num_pairs):
    unique_neighboring_pairs = set()

    # iterate through neighborhood, get number of unique pairs
    for ID_i, neighbors in neighborhood.items():
        for N_ID in neighbors:
            pair = (ID_i, N_ID)
            # if this pair isn't already in unique pairs (accounting for order not mattering)
            if (ID_i, N_ID) not in unique_neighboring_pairs \
                    and (N_ID, ID_i) not in unique_neighboring_pairs:
                # add the pair to unique pairs
                unique_neighboring_pairs.add(pair)

    num_neighboring_pairs = len(unique_neighboring_pairs)

    p_N = num_neighboring_pairs / num_pairs

    print("P(N) = # neighboring pairs / # pairs")
    print("=", num_neighboring_pairs, "/", num_pairs, "=", p_N)
    print()

    return p_N


# return bool whether T and N are dependent
# in temporal activation data
# (if True, T and N are dependent on one another)
def T_N_dependent(time_dep, neighborhood, num_time_dep_pairs):
    num_pairs = math.comb(len(chrom_IDs), 2)  # num chroms choose 2 = number of possible pairs

    print("T:", time_dep)
    print()
    print("N:", neighborhood)

    p_T_and_N = get_p_T_and_N(time_dep, neighborhood, num_pairs)  # P(T AND N)
    p_T = get_p_T(num_time_dep_pairs, num_pairs)  # P(T)
    p_N = get_p_N(neighborhood, num_pairs)  # P(N)

    # if P(T AND N) = P(T) * P(N), T and N are independent
    difference = abs(p_T_and_N - p_T*p_N)

    # print calculations for T and N being independent
    print("If T and N are independent, we expect P(T AND N) = P(T) * P(N)")
    print("P(T):", p_T)
    print("P(N):", p_N)
    print("P(T AND N):", p_T_and_N)
    print("P(T)*P(N):", p_T*p_N)
    print("Difference:", difference)
    print()

    if difference < 0.01:
        dep = False  # they are independent
    else:
        dep = True  # they are dependent

    return dep


# get the centroids from the spreadsheet, populates the centroids dict so we can use that instead
def extract_centroids():
    # load workbook from .xlsx file
    wb = openpyxl.load_workbook(filepath)
    # get centroids sheet
    centroids_sheet = wb["Centroids"]
    # initialize dictionary
    centroids = {}

    # Put the centroid of each chrom into the centroids dictionary
    for colNumber in range(2, centroids_sheet.max_column):
        cur_ID = centroids_sheet.cell(1, colNumber).value  # get ID num of cur chrom

        # get this chrom's centroid vals
        x = float(centroids_sheet.cell(2, colNumber).value)
        y = float(centroids_sheet.cell(3, colNumber).value)
        cur_centroid = [x, y]

        centroids[cur_ID] = cur_centroid  # populate centroid dict

    return centroids


# return ROI from parsing the xlsx filepath
def extract_ROI():
    # xlsx file name (with extension)
    filename_ext = os.path.basename(filepath)
    # xlsx file name without extension
    filename_no_ext = os.path.splitext(filename_ext)[0]
    # extract ROI from xlsx file name
    ROI = filename_no_ext.split("_")[1:]  # [x, y, w, h] of ROI
    # make ROI nums into actual nums (not strings)
    for i in range(len(ROI)):
        ROI[i] = int(ROI[i])

    return ROI


# return directory of context image that we'll be annotating
# later in the draw_pathway() function
def get_context_image_dir():
    ROI = extract_ROI()

    # TODO update vidprocessing.py to make centroids not relative to ROI ?

    # set directory of the context image (that we'll be drawing on)
    if args["arrows_img"] == "auto_detect":
        context_image_dir = os.path.splitext(filepath)[0] + "_ROI_context.png"
    else:
        context_image_dir = args["arrows_img"]

    # if the context image cannot be found, raise error
    if not os.path.exists(context_image_dir):
        raise Exception("No image found at " + context_image_dir + ". Please run vidProcessing.py "
                                                                "on the same video + ROI to "
                                                                "generate the context image. Also "
                                                                "ensure that the context image "
                                                                "has the aforementioned "
                                                                   "name/filepath.")

    return context_image_dir, ROI


# return matplotlib name of cmap based on selection from args
def get_cmap_name():
    if args["arrow_color"] == "rainbow_spectrum":
        return "Spectral"
    elif args["arrow_color"] == "cyan/magenta spectrum":
        return "cool"
    elif args["arrow_color"] == "blue/green spectrum":
        return "winter"


# generate discreet points of color along a spectrum
# so we can better distinguish different activations
# chronologically based on what color their arrow/label is
def get_spectrum_colors(num_activations, cmap_name):
    spectrum_colors = []

    x = np.linspace(0, num_activations - 1, num_activations)

    # define color map
    cmap = cm.get_cmap(cmap_name)
    # need to normalize because color maps are defined in [0, 1]
    norm = colors.Normalize(0, num_activations - 1)

    # fill in colors array with
    for i in x:
        color_i = cmap(norm(i))[:-1]
        r = int(color_i[0] * 255)
        g = int(color_i[1] * 255)
        b = int(color_i[2] * 255)
        spectrum_colors.append((b, g, r))

    # TODO (if time) -- check why spectrum_colors is normalizing to default len (50)
    return spectrum_colors


# draws a pathway of activation chronologically
# draws arrows on image from centroid to centroid
# takes in a list of activations sorted chronologically
# e.g [["C12", 23], ["C45", 28], ...]
# expects the context image to exist already from running vidProcessing.py
def draw_pathway(activations, centroids):
    # get path to image that we'll be annotating
    context_image_dir, ROI = get_context_image_dir()

    # open image
    image = cv2.imread(context_image_dir)

    # generate colors to display arrow/label colors from a spectrum (to distinguish chronologically)
    if args["arrow_color"] != "uniform (green/blue)":
        cmap_name = get_cmap_name()
        spectrum = get_spectrum_colors(len(activations), cmap_name)

    act_offset = 1  # adjustment to displayed activation number
    # if the user wants to only look at specific activations (e.g. 3rd through 6th)
    if args["activations_of_interest"] != "":
        act_offset = int(args["activations_of_interest"].split(" ")[0])  # 1st activation num of interest

    # draw arrow between each activation chronologically
    for i in range(len(activations) - 1):
        start_chrom = activations[i][0]  # ID of chrom where arrow will start from
        # adjust for ROI since centroid is relative to ROI, not full image
        start_x = int(centroids[start_chrom][0]) + ROI[0]
        start_y = int(centroids[start_chrom][1]) + ROI[1]

        end_chrom = activations[i+1][0]  # ID of chrom where arrow will end at
        end_x = int(centroids[end_chrom][0]) + ROI[0]
        end_y = int(centroids[end_chrom][1]) + ROI[1]

        act_num_label = str(i + act_offset)

        # set arrow and text colors
        if args["arrow_color"] != "uniform (green/blue)":  # color based on spectrum
            arrow_color = spectrum[i]
            text_color = spectrum[i]  # TODO consider changing to black text (0, 0, 0)
        else:  # default colors (green arrows, blue labels)
            arrow_color = (0, 255, 0)
            text_color = (255, 0, 0)

        # draw arrow
        cv2.arrowedLine(image, (start_x, start_y), (end_x, end_y), arrow_color, 1, cv2.LINE_AA,
                        0, 0.06)
        # add text labelling what number activation it is (1, 2, etc)
        cv2.putText(image, act_num_label, (int((start_x + end_x) / 2), int((start_y + end_y) / 2)),
                    cv2.FONT_HERSHEY_DUPLEX, 0.4, text_color, 1)

    # set filename/path for annotated image
    pathways_dir = context_image_dir.replace("_ROI_context.png", "_pathways.png")

    # save image
    cv2.imwrite(pathways_dir, image)

    print("Saved pathways image to " + pathways_dir)

    # display image
    cv2.imshow("pathways", image)
    cv2.waitKey(0)


# summary function of the program
def main():
    parse_args()  # parse input from user
    set_global_nums()  # set global nums based on user input

    # get activations from spreadsheet
    chronological_activations, activations_by_chrom = get_sorted_activations()

    # sort activation times into windows
    hist_act_df = make_act_histogram(activations_by_chrom)

    # find time-dependent pairs
    time_dep, num_time_dep_pairs = get_time_dep_pairs(hist_act_df)

    # get neighborhood from spreadsheet
    neighborhood = get_neighborhood()

    # calc if T and N are dependent
    T_N_dep = T_N_dependent(time_dep, neighborhood, num_time_dep_pairs)

    # display results
    if T_N_dep:
        print("T and N are dependent.")
    else:
        print("T and N are independent.")

    # get centroids from sheet
    centroids = extract_centroids()

    # if the user wants to only look at specific activations (e.g. 3rd through 6th)
    if args["activations_of_interest"] != "":
        act_start, act_end = args["activations_of_interest"].split(" ")

        # draw pathway of specific activations on image
        # note: + 2 to adjust for slicing being non-inclusive for end num and user
        # probably 1-indexing the activations
        draw_pathway(chronological_activations[int(act_start):int(act_end) + 2], centroids)
    else:
        # draw pathway of all activations on image
        draw_pathway(chronological_activations, centroids)

    # close open image windows
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()
