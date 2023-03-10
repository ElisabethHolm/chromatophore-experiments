# Code by Elisabeth Holm
# Analyzes raw area and centroid data, computing
# variances, neighbors, activation events, etc
# Python 3.10.4
# Feb 2023 - Present

import math
from collections import OrderedDict
import statistics
import openpyxl  # openpyxl 3.1.1
import os  # for checking if it's a .xls or .xlsx file
import pyexcel as p  # pyexcel 0.7.0, pyexcel-xls 0.7.0, pyexcel-xlsx 0.6.0, for converting .xls to .xlsx
from argparse import ArgumentDefaultsHelpFormatter, ArgumentParser # argparse 1.4.0
from gooey import Gooey, GooeyParser  # Gooey 1.0.8.1
from derivative import dxdt
import numpy as np

# 367 12 383 456  # ROI gilly chose

#filepath = "New_Vids/Sept_2022_Tests/TS-20220801155100769_391_61_363_354.xls"  # I chose this ROI, clear pathway, ROI near electrode
# later: use data from vid 6363 for all chroms going off at once
# 9896 - clear pathway
filepath = "New_Vids/Sept_2022_Tests/TS-20220801155100769_367_12_383_456.xls"

mm_nei_thresh = 3  # threshold distance (in mm) for neighbor # TODO play with thresh dist for neighbor

centroids = OrderedDict()  # dict with key: ID num, value: list [x,y] of its centroid coordinates
neighborhood = OrderedDict()  # dict with key: ID num, value: set of ID num(s) of its neighbors
chrom_IDs = []  # list with chrom IDs in order of how they appear on the spreadsheet, populate in main()

wb = 0  # global var, redefine to be workbook in main

args = []


@Gooey  # The GUI Decorator goes here
def parse_args():
    global args

    ap = GooeyParser(formatter_class=ArgumentDefaultsHelpFormatter,
                     conflict_handler='resolve')

    ap.add_argument("-s", "--spreadsheet", default=filepath, widget="FileChooser",
                    help="File path of the spreadsheet (.xls and .xlsx accepted)")
    ap.add_argument("-n", "--neighbor_radius", default=mm_nei_thresh, type=float,
                    help="Threshold for distance (in mm) of a neighbor")
    ap.add_argument("-a", "--act_deriv_thresh", default=0.0035, type=float,
                    help="Threshold value (of derivative of areas) for activation")
    # TODO find deact thresh val and set here
    ap.add_argument("-d", "--deact_deriv_thresh", default=-0.0035, type=float,
                    help="Threshold value (of derivative of areas) for activation")

    args = vars(ap.parse_args())


# set global nums based on args
def set_global_nums():
    global filepath
    global mm_nei_thresh

    filepath = args["spreadsheet"]
    mm_nei_thresh = args["neighbor_radius"]


# converts xls file to xlsx
def convert_xls_to_xlsx():
    global filepath

    p.save_book_as(file_name=filepath,
                   dest_file_name=filepath + "x")
    filepath = filepath + "x"


# get the centroids from the spreadsheet, populates the centroids dict so we can use that instead
def extract_centroids(centroids_sheet):
    global centroids
    global PIXELS_PER_MM

    # if using an older spreadsheet that doesn't have this data, use the most common val
    try:
        PIXELS_PER_MM = float(centroids_sheet.cell(10, 2).value)
        print(PIXELS_PER_MM)
    except:
        PIXELS_PER_MM = 58.3590729166667

    # Put the centroid of each chrom into the centroids dictionary
    for colNumber in range(2, centroids_sheet.max_column):
        cur_ID = centroids_sheet.cell(1, colNumber).value  # get ID num of cur chrom

        # get this chrom's centroid vals
        x = float(centroids_sheet.cell(2, colNumber).value)
        y = float(centroids_sheet.cell(3, colNumber).value)
        cur_centroid = [x, y]

        centroids[cur_ID] = cur_centroid  # populate centroid dict
        neighborhood[cur_ID] = set()  # prepare neighborhood dict to be filled later


# go through centroids sheet and determine which chroms are adjacent to each other
# based on a threshold Euclidean distance
def determine_neighbors():
    global neighborhood
    global mm_nei_thresh

    ID_nums = list(centroids.keys())
    cents = list(centroids.values())

    # iterate through each chrom and find all its neighbors
    for i in range(len(ID_nums)):
        ID_num = ID_nums[i]
        centroid = cents[i]
        # check against unchecked pns
        for j in (range(i + 1, len(ID_nums))):  # pn = potential neighbor
            pn_ID = ID_nums[j]
            pn_centroid = cents[j]
            # if a pn is close enough to the current chrom, count it as a neighbor
            if (math.dist(centroid, pn_centroid)/PIXELS_PER_MM) < mm_nei_thresh:
                # both chroms as neighbors of each other
                (neighborhood[ID_num]).add(pn_ID)
                neighborhood[pn_ID].add(ID_num)


# filter out any non-numbers from a list of data
# return a list of just the numerical data, and a list of which frames those data points came from
def get_only_nums(full_list):
    just_nums = []
    corresponding_frames = []
    for i in range(len(full_list)):
        if (type(full_list[i]) == float) or (type(full_list[i]) == int):
            just_nums.append(full_list[i])
            corresponding_frames.append(i)

    return just_nums, corresponding_frames


# compute the rolling variance time series for a single chrom
# if unable to compute variance (e.g not enough data points in window), say var = 0
# params: chromatophore ID num (e.g "C12"), the entire column of areas data (as a list) for chrom
def compute_var_data(chrom_ID, area_data):
    var_data = {}
    rolling_window_size = 20

    # generate variance data point for each
    for i in range(len(area_data)):
        # fill in first window-size frames as having a previous window variance of 0 (since there
        # aren't enough points to compute a variance
        if i < rolling_window_size - 1:
            window_var = 0
        else:
            # take window of data as the previous [window size] points before the ith point
            window_data = area_data[i - rolling_window_size: i - 1]

            # get only actual numbers from window data
            window_data, corresponding_frames = get_only_nums(window_data)

            if len(window_data) > 1:  # check that we have at least 2 data points, compute variance
                window_var = statistics.variance(window_data)  # rolling variance
                window_avg = statistics.mean(window_data)  # rolling avg
            else:  # if the window covers a gap in the data, put the variance as 0
                window_var = 0

        var_data[i] = window_var

    return var_data


# TODO, NOT YET IMPLEMENTED -- might not need to bc of size range method of computing activations
# determine activation time(s) for a single chrom using the variance time series
# param: the entire column of variation data (as a dict of frame_num:var)) for that chromatophore
def find_activations_via_var(chrom_ID, var_data):
    # iterate through each variance data point
    for frame_num in range(len(var_data.keys())):
        var_pt = var_data[frame_num]

        # TODO?
        # take avg slope between prev 3 data points -- testing frame 4, we'd avg slopes between
        # 1/2 2/3 3/4
        # if that avg slope exceeds 500 (or some threshold amount) then that counts as an event


# fill in gaps in data (essentially drawing a straight line between points across the gaps
def fill_in_data_gaps(data):
    filled_data = data.copy()
    just_nums, just_nums_frames = get_only_nums(data)
    start_gap_i = 0  # first frame without a num
    end_gap_i = 0  # last name without a num

    # iterate through all the actual numerical data
    for i in range(len(just_nums_frames)):
        # if this is the first one in the just_nums_frames list
        if i == 0:
            # if there is a gap that at the start of the data, hold the first real num constant
            if just_nums_frames[i] != 0:
                for j in range(just_nums_frames[i]):
                    filled_data[j] = just_nums[i + 1]

        # if this is the last one in the just_num_frames list
        elif i == len(just_nums_frames) - 1:
            # if there is a gap that lasts until the end of the vid, hold the last real num constant
            if just_nums_frames[i] != len(data) - 1:
                for j in range(just_nums_frames[i], len(data)):
                    filled_data[j] = just_nums[i]
        else:
            cur_existing_f = just_nums_frames[i]  # existing frame num
            next_existing_f = just_nums_frames[i + 1]  # existing frame num

            # if there is a gap that starts at the cur existing frame
            if cur_existing_f + 1 != next_existing_f:
                # calculate number of blank frames (between the two existing frames)
                num_blank_frames = next_existing_f - cur_existing_f - 1
                # create filler datapoints (note: endpoints = actual data)
                filler_points = np.linspace(data[cur_existing_f], data[next_existing_f], num=num_blank_frames + 2)
                # fill in missing data in data array
                for j in range(cur_existing_f, next_existing_f + 1):
                    filled_data[j] = filler_points.item(j - cur_existing_f)

    return filled_data


# returns all active frame #s (frames when chrom is active) and
# event frame #s (first time a chrom is activated and deactivated)
# using the derivative method for finding the activation and deactivation frames
def find_activations_via_deriv(chrom_ID, area_data):
    filled_area_data = fill_in_data_gaps(area_data)
    area_data_np = np.array(filled_area_data)
    frames_np = np.array(range(len(area_data)))
    #print(area_data_np)
    #print(frames_np)
    deriv = dxdt(area_data_np, frames_np, kind="finite_difference", k=1)
    # plot deriv and area data -- TODO take out once func fully tested
    # if chrom_ID == "C16":
    #     import matplotlib.pyplot as plt
    #     plt.plot(frames_np, deriv, label="deriv", marker='o', markersize=3)
    #     plt.plot(frames_np, filled_area_data, label="areas", marker='o', markersize=3)
    #     plt.legend()
    #     plt.show()


    active_frames = []
    act_event_frames = []
    deact_event_frames = []

    act_deriv_thresh = args['act_deriv_thresh']
    deact_deriv_thresh = args['deact_deriv_thresh']

    for i in range(len(deriv)):
        # if chrom is active in this frame
        if deriv[i] >= act_deriv_thresh or deriv[i] <= deact_deriv_thresh:
            active_frames.append(i)
            # if this is a frame when it switches from inactive to active
            if (i > 0 and (i-1 not in active_frames)) or i == 0:
                act_event_frames.append(i)
            # if it's still active
            if i == len(area_data) - 1:
                deact_event_frames.append("active at end of vid")

        # if this is a frame when it switches from active to inactive
        elif i > 0 and (i-1 in active_frames):
            deact_event_frames.append(i)

    return active_frames, act_event_frames, deact_event_frames


# returns all active frame #s (frames when chrom is active) and
# event frame #s (first time a chrom is activated and deactivated)
# using the % of size range method for finding the activation and deactivation frames
def find_activations_via_size_range(chrom_ID, area_data):
    just_nums, just_nums_frames = get_only_nums(area_data)

    min_a = min(just_nums)
    max_a = max(just_nums_frames)
    size_range = max_a - min_a
    percent_active_thresh = .05  # once chrom has reached this percent of its size range, count as active
    thresh_area = size_range * percent_active_thresh

    active_frames = []
    act_event_frames = []
    deact_event_frames = []

    for i in range(len(area_data)):
        # skip empty spreadsheet cells
        if (type(area_data[i]) != float) and (type(area_data[i]) != int):
            # if it loses the chrom, count that as deactivation # TODO check if this is what we want to do
            if i-1 in active_frames:
                deact_event_frames.append(i)
            continue

        # if chrom is active in this frame
        if (area_data[i] - min_a) >= thresh_area:
            active_frames.append(i)
            # if this is a frame when it switches from inactive to active
            if (i > 0 and (i-1 not in active_frames)) or i == 0:
                act_event_frames.append(i)
            # if it's still active
            if i == len(area_data) - 1:
                deact_event_frames.append("active at end of vid")

        # if this is a frame when it switches from active to inactive
        elif i > 0 and (i-1 in active_frames):
            deact_event_frames.append(i)

    return active_frames, act_event_frames, deact_event_frames


# return the data from a column in the sheet as a list
def get_col_as_list(sheet, col_num):
    col_data = []

    for i in range(1, sheet.max_row + 1):
        col_data.append(sheet.cell(row=i, column=col_num).value)

    return col_data


# for every chrom, determine its variance data and times of activation
def compute_var_and_activations():
    all_var_data = OrderedDict() # dict with key: chrom ID, value: that chrom's variance time series (a dict)
    all_activation_data = OrderedDict() # dict with key: chrom ID, value: [act_event_frames, deact_event_frames]
    areas_sheet = wb["Areas"]

    for colNumber in range(2, areas_sheet.max_column):
        chrom_col = get_col_as_list(areas_sheet, colNumber)  # get column as a list
        chrom_ID = chrom_col[0]

        var_data = compute_var_data(chrom_ID, chrom_col[1:])  # create var time series for chrom
        all_var_data[chrom_ID] = var_data

        # TODO not yet implemented -- may not need bc of deriv and size_range method
        #active_frames, act_event_frames, deact_event_frames = find_activations_via_var(chrom_ID, var_data)

        active_frames, act_event_frames, deact_event_frames = find_activations_via_deriv(chrom_ID, chrom_col[1:])

        # using the % of size range method for finding the activation and deactivation frames
        #active_frames, act_event_frames, deact_event_frames = find_activations_via_size_range(chrom_ID, chrom_col[1:])
        all_activation_data[chrom_ID] = [act_event_frames, deact_event_frames]

    return all_var_data, all_activation_data


# creates a new sheet in the wb
# AND deletes the old sheet if there's one of the same name in that wb already
# returns new sheet
def create_new_sheet(sheet_name):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    return wb.create_sheet(sheet_name)


# take in a list of neighbors, sort by their chromatophore ID (numerical)
def sort_neighbors(n_set):
    neighbors = list(n_set)
    # separate out just the ID nums, sort them numerically
    just_nums = []
    for chrom_ID in neighbors:
        just_nums.append(int(chrom_ID[1:]))
    just_nums = sorted(just_nums)

    # add back in the "C" before each ID num, forming the now-sorted list
    sorted_neighbors = []
    for ID_num in just_nums:
        sorted_neighbors.append("C" + str(ID_num))

    return sorted_neighbors


def save_neighbors_data():
    nei_sheet = create_new_sheet('Neighbors')

    nei_sheet.cell(row=1, column=1).value = "Chromatophore ID"

    curRow = 2
    max_neighbors = 0

    # for every chromatophore, add each of its neighbors to the spreadsheet
    for chrom_ID, neighbors in neighborhood.items():
        # add chrom_ID label
        nei_sheet.cell(row=curRow, column=1).value = chrom_ID

        # add neighbors to sheet
        neighbors_l = sort_neighbors(neighbors)
        for i in range(len(neighbors_l)):
            nei_sheet.cell(row=curRow, column=i + 2).value = neighbors_l[i]

        # keep track of max number of neighbors so we know how many N1, N2,... labels to put
        if len(neighbors_l) > max_neighbors: max_neighbors = len(neighbors_l)

        curRow += 1

    # add N1, N2,... labels at top
    for i in range(2, 2 + max_neighbors):
        nei_sheet.cell(row=1, column=i).value = "N" + str(i-1)


# save the variance data to a new sheet in the spreadsheet
def save_variance_data(all_var_data):
    varSheet = create_new_sheet('Variances')

    varSheet.cell(row=1, column=1).value = "frame"

    # add column in the first column that says each frame number
    for i in range(2, wb["Areas"].max_row + 1):
        varSheet.cell(row=i, column=1).value = i - 2

    curCol = 2

    # iterate through each identified chromatophore
    for id_num, var_data in all_var_data.items():
        # add ID num on row 0
        varSheet.cell(row=1, column=curCol).value = id_num

        # add var time series in each column
        for frame_num, var in var_data.items():
            varSheet.cell(row=frame_num + 2, column=curCol).value = var

        curCol += 1


# save all activation data to the spreadsheet
def save_activation_data(all_activation_data):
    act_sheet = create_new_sheet('Activations')

    act_sheet.cell(row=1, column=1).value = "Chromatophore ID"

    curRow = 2
    max_events = 0
    # for every chromatophore, add each of its activations and deactivations to the spreadsheet
    for chrom_ID, act_data in all_activation_data.items():
        act_event_frames = act_data[0]
        deact_event_frames = act_data[1]

        act_sheet.cell(row=curRow, column=1).value = chrom_ID

        # add event data into spreadsheet
        for i in range(len(act_event_frames)):
            act_sheet.cell(row=curRow, column=2 + i*3).value = act_event_frames[i]  # activation
            act_sheet.cell(row=curRow, column=3 + i*3).value = deact_event_frames[i]  # deactivation
            if type(deact_event_frames[i]) != str:
                act_sheet.cell(row=curRow, column=4 + i*3).value = deact_event_frames[i] - act_event_frames[i]  # duration of activation
            else:
                act_sheet.cell(row=curRow, column=4 + i * 3).value = deact_event_frames[i]  # "active at end of vid" --> duration can't be computed

        # keep track of max number of events so we know how many A1, D1, Dur1,... labels to put
        if len(act_event_frames) > max_events: max_events = len(act_event_frames)

        curRow += 1

    # add A1, D1, Dur1,... labels at top
    for i in range(max_events):
        act_sheet.cell(row=1, column=2 + i*3).value = "A" + str(i+1)  # activation
        act_sheet.cell(row=1, column=3 + i*3).value = "D" + str(i+1)  # deactivation
        act_sheet.cell(row=1, column=4 + i*3).value = "Dur" + str(i+1)  # duration


#def identify_pathways():
    #print("wahoo")

#def predict_independence():
    #print("bat time")


def main():
    global wb
    global chrom_IDs

    # adjust for user-selected options
    parse_args()

    set_global_nums()

    file_ext = os.path.splitext(filepath)[1]

    # convert .xls to .xlsx (so we can use openpyxl) if the input file is .xls
    if file_ext == ".xls":
        convert_xls_to_xlsx()
    elif file_ext != ".xlsx":
        raise Exception("Invalid input file. Must be either .xls or .xlsx")

    # load workbook
    wb = openpyxl.load_workbook(filepath)

    # get all chrom IDs in order, save to chrom_IDs
    for cell in wb["Areas"][1]:
        if cell != wb["Areas"].cell(1, 1):
            chrom_IDs.append(cell.value)

    # extract data from centroids sheet, populate centroids dict
    extract_centroids(wb["Centroids"])

    # determine which chroms are neighbors with which, populate neighborhood dict
    determine_neighbors()
    # save neighbors data
    save_neighbors_data()

    # compute variance and activation data, then save to the spreadsheet
    all_var_data, all_activation_data = compute_var_and_activations()
    save_variance_data(all_var_data)
    save_activation_data(all_activation_data)

    # save the file (under the same name, thus replacing the unedited file)
    wb.save(filepath)

    print("Saved updated spreadsheet to " + filepath)

if __name__ == "__main__":
    main()

'''
reader = csv.DictReader(open('bats.csv'))

sum_G1 = 0
sum_G2 = 0
sum_G3 = 0
sum_G4 = 0
sum_G5 = 0
sum_T = 0
num_bats = 0

sum_G1_given_T = 0
sum_G2_given_T = 0
sum_G3_given_T = 0
sum_G4_given_T = 0
sum_G5_given_T = 0

for bat in reader:
    num_bats += 1

    if bat['G1'].lower() == "true":
        sum_G1 += 1
    if bat['G2'].lower() == "true":
        sum_G2 += 1
    if bat['G3'].lower() == "true":
        sum_G3 += 1
    if bat['G4'].lower() == "true":
        sum_G4 += 1
    if bat['G5'].lower() == "true":
        sum_G5 += 1
    if bat['T'].lower() == "true":
        sum_T += 1
        if bat['G1'].lower() == "true":
            sum_G1_given_T += 1
        if bat['G2'].lower() == "true":
            sum_G2_given_T += 1
        if bat['G3'].lower() == "true":
            sum_G3_given_T += 1
        if bat['G4'].lower() == "true":
            sum_G4_given_T += 1
        if bat['G5'].lower() == "true":
            sum_G5_given_T += 1

prob_G1 = sum_G1/num_bats
prob_G2 = sum_G2/num_bats
prob_G3 = sum_G3/num_bats
prob_G4 = sum_G4/num_bats
prob_G5 = sum_G5/num_bats
prob_T = sum_T/num_bats

print("P(GX) -- Probability of each event independently happening (individual events/total num of bats)")
print("G1: " + str(prob_G1))
print("G2: " + str(prob_G2))
print("G3: " + str(prob_G3))
print("G4: " + str(prob_G4))
print("G5: " + str(prob_G5))
print("T:  " + str(prob_T))
print()

p_G1_given_T = sum_G1_given_T/num_bats
p_G2_given_T = sum_G2_given_T/num_bats
p_G3_given_T = sum_G3_given_T/num_bats
p_G4_given_T = sum_G4_given_T/num_bats
p_G5_given_T = sum_G5_given_T/num_bats

print("P(GX|T) -- Probability of GX appearing given T is true (sum_GX_given_T/num_bats)")
print("G1|T: " + str(p_G1_given_T))
print("G2|T: " + str(p_G2_given_T))
print("G3|T: " + str(p_G3_given_T))
print("G4|T: " + str(p_G4_given_T))
print("G5|T: " + str(p_G5_given_T))
print()

def predict_indep(p_GX, p_GX_given_T):
    difference = abs(p_GX*prob_T - p_GX_given_T)
    print("Difference:" + str(difference))
    if difference < 0.01 :
        print("Result: NOT independent of T")
    else:
        print("Result: independent of T")
    print()

print("Prediction of independence based on difference between P(GX)*P(T) and P(GX|T) (dependent if <0.01 apart from each other)")
print("G1 :")
predict_indep(prob_G1, p_G1_given_T)
print("G2 :")
predict_indep(prob_G2, p_G2_given_T)
print("G3 :")
predict_indep(prob_G3, p_G3_given_T)
print("G4 :")
predict_indep(prob_G4, p_G4_given_T)
print("G5 :")
predict_indep(prob_G5, p_G5_given_T)
'''
