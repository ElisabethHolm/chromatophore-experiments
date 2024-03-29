# Code by Elisabeth Holm
# Calculates the changing areas of chromatophores from raw video data
# using computer vision (openCV)
# Program 1/3 in pipeline
# Python 3.10.4
# April 2022 - Present

# so I don't have to type it out every time:
# cd Downloads/Coding/Chromatophore_Research/dataProcessing
# python3 vidProcessing.py

# import the necessary packages
import imutils  # imutils 0.5.4
import cv2  # opencv-python 4.6.0.66
import numpy as np  # numpy 1.22.4
from scipy.spatial import distance as dist  # scipy 1.8.1
from collections import OrderedDict
import matplotlib.pyplot as plt  # matplotlib 3.5.2
from openpyxl.workbook import Workbook  # openpyxl 3.1.1
import math
from argparse import ArgumentDefaultsHelpFormatter, ArgumentParser # argparse 1.4.0
from gooey import Gooey, GooeyParser  # Gooey 1.0.8.1
import os


# these values will be adjusted in setGlobalNums() based on cmd arguments
DISPLAY_SCALE = 0.5
(ORIGINAL_WIDTH, ORIGINAL_HEIGHT) = (1920, 1080)
CHROM_PIXEL_DIAMETER = 0
PIXELS_PER_MM = 0
MAX_DISAPPEARED = 30

nextObjectID = 0
disappeared = OrderedDict()  # objectID : num of frames it has disappeared for
matchedCentroids = OrderedDict()  # objectID : centroid
chromAreas = OrderedDict()  # objectID : area time series (frame : area)
initialCentroids = OrderedDict()  # object ID : centroid when it was first detected

# directory of raw video to process (relative to where this program file is)
#curVidPath = "15_41.811_12TEA+TTX.m4v"  # oldest vid, good w option B or C
#curVidPath = "New_Vids/April_22/VID00062.AVI"  # main test -- new, no muscle, great w/ option C (outlines clear but hollow)
#curVidPath = "New_Vids/April_22/VID00062_cropped.mov"  # smaller portion of frame from vid in line above, used for testing program on small scale
#curVidPath = "New_Vids/April_22/45_4-20-22.AVI" # new, decent w/ option B (reduced noise but can't identify big one), or E with (7, 7)
#curVidPath = "New_Vids/April_22/VID00065.AVI"  # pulled from Puneeta on 6/23 (not from 6/23 experiment)
#curVidPath = "New_Vids/April_22/VID00056.AVI"  # pulled from Puneeta on 6/23 (not from 6/23 experiment), struggles w keeping ID nums
#curVidPath = "New_Vids/June_22/TS-20220630153941965.avi"  # from June lab visit
curVidPath = "New_Vids/Sept_2022_Tests/TS-20220801155100769.mp4"

vidName_ext = ""  # video file name (with extension)
vidName_no_ext = ""  # video file name without extension

args = []

@Gooey # The GUI Decorator goes here
def parse_args():
	global args
	ap = GooeyParser(formatter_class=ArgumentDefaultsHelpFormatter,
								 conflict_handler='resolve')
	ap.add_argument("-r", "--ROI", type=str,
					help="x, y, w, h of region of interest, each number separated by a space. "
						 "Leave blank to select ROI by drawing a box on the frame.")
	ap.add_argument("-v", "--vid", default=curVidPath, widget="FileChooser",
					help="File path of the video (.mp4 and .avi accepted)")
	ap.add_argument("-s", "--stop", default="off", choices=["off", "on"],
					help="If you want the video to stop and wait for a space bar at every frame")
	ap.add_argument("-d", "--display_vid", default="on", choices=["off", "on"],
					help="If you want the video to display as the program runs")
	ap.add_argument("-f", "--frame_cap", default="off", choices=["off", "on, save ID frame", "on, save entire process"],
					help="If you want to save every frame to a folder (in the same location as your "
						 "video file). Helpful for comparing "
						 "individual frames")
	ap.add_argument("-o", "--watch_only", default="off", choices=["off", "on"],
					help="If you want to only re-watch the video (e.g. to check if the ID numbers "
						 "change over time), but not save any spreadsheet data")
	ap.add_argument("-j", "--jellyfish", default="off",
					choices=["on",
							 "off"],
					help="Adjusts program to run for jellyfish vids. Note: overrides \"prioritize\" setting.")
	ap.add_argument("-p", "--prioritize", default="more accurate IDs and area",
					choices=["more accurate IDs and area",
							 "prevent merging of adjacent chroms", "manual"],
					help="Adjusts parameters of detection to priortize certain features -- "
						 "IDs (best for well-spaced chromatophores), merging (best if many chroms "
						 "overlap when expanded), manual (set num manually in \"manual\")")
	ap.add_argument("-a", "--manual", default=3.0, type=float,
					help="Manually adjusts parameters of detection (if \"prioritize\" set to "
						 "\"manual\"). Scalar multiplied by PIXELS_PER_MM to set CHROM_PIXEL_DIAMETER. "
						 "Should be ~the avg mm diameter of the object (~3.0 for chroms).")
	ap.add_argument("-c", "--save_ROI_context_img", default="on", choices=["off", "on"],
					help="Save the image that shows where the ROI is in the uncropped frame")
	ap.add_argument("-m", "--magnification", default=7, type=float,
					help="Magnification level of microscope when video was filmed")
	ap.add_argument("-e", "--display_scale", default=DISPLAY_SCALE, type=float,
					help="Scalar for size that frame is displayed")
	ap.add_argument("-i", "--ID_draw_shape", default="contours", choices=["contours", "rotated rectangles"],
					help="Shape that outlines each chromatophore in the displayed ID frame (does not "
						 "affect the area calculation)")

	args = vars(ap.parse_args())

# sets the global constants for
# window width and chromatophore pixel diameter based on
# command line arguments
def setGlobalNums():
	global DISPLAY_SCALE
	global CHROM_PIXEL_DIAMETER
	global PIXELS_PER_MM
	global ORIGINAL_WIDTH
	global ORIGINAL_HEIGHT
	global curVidPath
	global args
	global vidName_ext
	global vidName_no_ext

	DISPLAY_SCALE = float(args["display_scale"])
	magnification = float(args["magnification"])
	curVidPath = args["vid"]

	# open the video
	vid = cv2.VideoCapture(curVidPath)
	# read current frame from the video
	ok, frame = vid.read()
	# stop if error reading frame or end of vid reached
	if not ok:
		raise Exception("Uh oh! There was a problem opening the video. Check that the file exists, "
						"has read permissions on, and that the file path fed into the program is "
						"\correct.")

	ORIGINAL_HEIGHT, ORIGINAL_WIDTH, color_channels = frame.shape   # get image info from 1st frame

	# video file name (with extension)
	vidName_ext = os.path.basename(curVidPath)
	# video file name without extension
	vidName_no_ext = os.path.splitext(vidName_ext)[0]

	if os.path.splitext(vidName_ext)[1].lower() not in [".mp4", ".avi"]:
		raise Exception("Invalid input file. Must be either .mp4 or .avi")

	# settings for chromatophore mode
	if args["jellyfish"] == "off":
		# pixels per mm adjusting for magnification, original resolution, and window width
		# Note: at x50 magnification, number of pixels for 1 mm is 800.353 on 1920x1080 display
		#( 800.353 pixels per mm when at magnification 50) * (ratio of current magnification to mag 50) *
		# (ratio of video's resolution to resolution that calibration measurements were taken on)
		PIXELS_PER_MM = (800.353 / 1) * (magnification / 50) * (ORIGINAL_WIDTH / 1920)

		# Note: a fully expanded market squid chromatophore is ~1.7 mm in diameter
		# Note 2: jellyfish mode overrides the "prioritize" setting
		if (args["prioritize"] == "more accurate IDs and area"):
			CHROM_PIXEL_DIAMETER = int(PIXELS_PER_MM * 3)
		elif (args["prioritize"] == "prevent merging of adjacent chroms"):
			CHROM_PIXEL_DIAMETER = int(PIXELS_PER_MM * 2.9)
		elif (args["prioritize"] == "manual"):
			CHROM_PIXEL_DIAMETER = int(PIXELS_PER_MM * args["manual"])
			
	else:  # settings for jellyfish mode
		PIXELS_PER_MM = (800.353 / 1) * (magnification / 50) * (ORIGINAL_WIDTH / 1920)
		CHROM_PIXEL_DIAMETER = int(PIXELS_PER_MM * 3)
		# opt 1: * 30
		# opt 2: * 3
	
	# ensure CHROM_PIXEL_DIAMETER is odd (requirement for param it's passed into)
	if CHROM_PIXEL_DIAMETER % 2 == 0:
		CHROM_PIXEL_DIAMETER += 1


# Set ROI from command line or the user manually choosing
# takes in the first frame of the video and returns ROI coordinates/dimensions
def getROI(frame):
	# if an ROI was passed in as a command line arg, use that ROI
	if args["ROI"] != None:
		[ROI_x, ROI_y, ROI_width, ROI_height] = args["ROI"].split(" ")
		ROI_x = int(ROI_x)
		ROI_y = int(ROI_y)
		ROI_width = int(ROI_width)
		ROI_height = int(ROI_height)

	# Have user manually choose ROI, record (x,y) and height/width
	else:
		# resize frame before displaying to choose ROI from
		display_frame = imutils.resize(frame.copy(), width=(int(DISPLAY_SCALE * ORIGINAL_WIDTH)))
		[ROI_x, ROI_y, ROI_width, ROI_height] = cv2.selectROI("Select ROI", display_frame,
															  fromCenter=False,showCrosshair=True)
		cv2.destroyWindow("Select ROI")

		# rescale ROI for original frame
		[ROI_x, ROI_y, ROI_width, ROI_height] = [int(ROI_x/DISPLAY_SCALE), int(ROI_y/DISPLAY_SCALE),
												 int(ROI_width/DISPLAY_SCALE), int(ROI_height/DISPLAY_SCALE)]

	print("Region selected:")
	print("x:", ROI_x, " y:", ROI_y, " width:", ROI_width, " height:", ROI_height)

	return ROI_x, ROI_y, ROI_width, ROI_height


# masks the current frame from the video to identify only the chromatophores
# returns the masked frame
def maskChromatophores(curFrame):
	# blue, green, red = cv2.split(curFrame)  # could use for doing things based on color

	# convert frame to grayscale
	gray_frame = cv2.cvtColor(curFrame, cv2.COLOR_BGR2GRAY)
	# apply Gaussian blur to reduce noise
	if args["jellyfish"] == "off":
		gray_frame = cv2.GaussianBlur(gray_frame, (7, 7), 0)

		# identifies darkest parts of frame in order to mask out chromatophores from background
		#threshold = cv2.adaptiveThreshold(gray_frame, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 5, 2)  # B. good, semi noisy/inaccurate blur (11, 11) !
		#threshold = cv2.adaptiveThreshold(gray_frame, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 5, 3)  # C. some noise but true to size, blur (3, 3) !!
		#threshold = cv2.adaptiveThreshold(gray_frame,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY_INV, 3 ,1)  # D. higher threshhold, noisy but true to size, blur (13, 13)
		#threshold = cv2.adaptiveThreshold(gray_frame, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 3, 2)  # E. not super solid but good mask, blur (5, 5)
		threshold = cv2.adaptiveThreshold(gray_frame, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, CHROM_PIXEL_DIAMETER, 10)  # currently using
		# 2nd to last param should be roughly the num of pixels of the diameter of an avg chromatophore in the frame

		'''
		# convert to LAB color space
		rgb_img = cv2.cvtColor(gray_frame, cv2.COLOR_GRAY2RGB)  # gray to RGB
		bgr_img = cv2.cvtColor(rgb_img, cv2.COLOR_RGB2BGR) # RGB to GBR
		lab = cv2.cvtColor(bgr_img, cv2.COLOR_BGR2LAB)  # BGR to lab
		# get luminance channel
		l_component = lab[:, :, 0]
		# setting threshold level at 110, 125 also p good
		ret, threshold = cv2.threshold(l_component, 215, 255, cv2.THRESH_BINARY_INV)
		'''
	else: 	# jellyfish settings
		gray_frame = cv2.GaussianBlur(gray_frame, (7, 7), 0)
		threshold = cv2.adaptiveThreshold(gray_frame, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, CHROM_PIXEL_DIAMETER, 2) 
		 # opt 1: last param 10
		 # opt 2: last param 2

	return gray_frame, threshold  # return gray/blurred and masked frame


# defines and draws the contour lines based on the masked frame
# takes in the binary masked image and original image, returns the list of chromatophores
def contour(threshold, origImg):
	# detect the contours on the binary image
	contours, hierarchy = cv2.findContours(image=threshold, mode=cv2.RETR_EXTERNAL, method=cv2.CHAIN_APPROX_NONE)

	# only take 3 biggest contours bc only 1 jellyfish but might accidentally contour wall or electrode
	if args["jellyfish"] == "on":
		contours = sorted(contours, key=cv2.contourArea, reverse=True)[:3]

	# draw contours on the original image
	contourCopy = origImg.copy()
	cv2.drawContours(image=contourCopy, contours=contours, contourIdx=-1, color=(0, 255, 0), thickness=1)

	return contourCopy, contours


# creates rotated bounding rectangles for each contour in contours
# returns a list of bounding rects
def createRotatedBRects(contours):
	return [cv2.minAreaRect(c) for c in contours]


# draws rotated rectangle
# takes in a rotated rectangle and image to draw it on
# returns the image with the rectangle drawn on it
def drawRotatedRect(rect, img):
	# calculate corners of box
	box = cv2.boxPoints(rect)
	box = np.int0(box)

	# draw contour of box on image
	cv2.drawContours(img, [box], 0, (0, 0, 255), 1)

	return img


# sorts the given contours from left to right
# returns the sorted contours and sorted bounding boxes
def sortLR(contours):
	# construct the list of bounding boxes
	boundingBoxes = createRotatedBRects(contours)

	# sort from left to right
	(sortedContours, sortedBoundingBoxes) = zip(*sorted(zip(contours, boundingBoxes),
		key=lambda b:b[1][0][0]))

	# return the list of sorted contours and bounding boxes
	return sortedContours, sortedBoundingBoxes


# sorts the given contours from top to bottom
# returns the sorted contours and bounding boxes
def sortTB(contours):
	# construct the list of bounding boxes
	boundingBoxes = createRotatedBRects(contours)

	# sort from top to bottom
	(sortedContours, sortedBoundingBoxes) = zip(*sorted(zip(contours, boundingBoxes),
		key=lambda b:b[1][0][1]))

	# return the list of sorted contours and list of sorted bounding boxes
	return sortedContours, sortedBoundingBoxes


# takes in the chromatophore contours found in the first frame (ff)
# of the video; Builds the chromAreas list of chromatophores
# (setting up to later add area data for each chromatophore)
def createAreasDict(ff_contours):
	global chromAreas

	# create "new chromatophore" in chromAreas for every chromatophore in 1st frame
	for i in range(0, len(ff_contours)):
		# create new chromatophore in chromAreas
		chromAreas[i] = {}
		i += 1


# populate the matchedCentroids dictionary with (objectID: centroid)
# in the first frame of the video when there is
# no previous frame to compare to
def createMatchedCentroids(curCentroids):
	for centroid in curCentroids:
		register(centroid)


# put the centroids of each bounding box in a
# list with the center of each bounding box in the
# same index as the original bounding box
# takes in list of bounding boxes and contours
# note: index of bBoxes and contours must correspond
# returns centroids and a dict that maps each centroid to its contour
def getCentroids(bBoxes, contours):
	centroids = []
	centroidsToContours = {}

	for curIndex in range(0, len(bBoxes)):
		centroid = bBoxes[curIndex][0]
		centroids.append(centroid)
		centroidsToContours[centroid] = contours[curIndex]

	return centroids, centroidsToContours


# register a new chromatophore
def register(centroid):
	global matchedCentroids
	global disappeared
	global nextObjectID
	global chromAreas

	# when registering an object we use the next available object
	# ID to store the centroid
	matchedCentroids[nextObjectID] = centroid
	disappeared[nextObjectID] = 0
	chromAreas[nextObjectID] = {}
	# Note: + ROI_x and ROI_y make centroids relative to full image, not ROI, for when saving to
	# spreadsheet. Throughout the program they are relative to ROI, so they will appear as different
	# nums that refer to the same place on the skin sample
	initialCentroids[nextObjectID] = (centroid[0] + ROI_x, centroid[1] + ROI_y)
	nextObjectID += 1


# deregister given chromatophore in matchedCentroids and disappeared dicts
def deregister(objectID):
	global matchedCentroids
	global disappeared

	# to deregister an object ID we delete the object ID from
	# our respective dictionaries
	del matchedCentroids[objectID]
	del disappeared[objectID]


# compare current centroids with previous centroids to
# determine the closest matches
# and return the calculated IDs of each current centroid.
# takes in the list of current centroids and dict of previous matched centroids
# returns dict of ID_num:current centroids matched
# code based on https://pyimagesearch.com/2018/07/23/simple-object-tracking-with-opencv/
def matchCentroids(curCentroids):
	global matchedCentroids
	global nextObjectID
	global disappeared

	# grab the set of object IDs and corresponding centroids
	objectIDs = list(matchedCentroids.keys())
	objectCentroids = list(matchedCentroids.values())

	# compute the distance between each pair of object
	# centroids and input centroids, respectively -- our
	# goal will be to match an input centroid to an existing
	# object centroid
	D = dist.cdist(np.array(objectCentroids), curCentroids)

	# in order to perform this matching we must (1) find the
	# smallest value in each row and then (2) sort the row
	# indexes based on their minimum values so that the row
	# with the smallest value is at the *front* of the index
	# list
	rows = D.min(axis=1).argsort()

	# next, we perform a similar process on the columns by
	# finding the smallest value in each column and then
	# sorting using the previously computed row index list
	cols = D.argmin(axis=1)[rows]

	# in order to determine if we need to update, register,
	# or deregister an object we need to keep track of which
	# of the rows and column indexes we have already examined
	usedRows = set()
	usedCols = set()
	# loop over the combination of the (row, column) index
	# tuples
	for (row, col) in zip(rows, cols):
		# if we have already examined either the row or
		# column value before, ignore it
		# val
		if row in usedRows or col in usedCols:
			continue
		# otherwise, grab the object ID for the current row,
		# set its new centroid, and reset the disappeared
		# counter
		objectID = objectIDs[row]
		matchedCentroids[objectID] = curCentroids[col]
		disappeared[objectID] = 0
		# indicate that we have examined each of the row and
		# column indexes, respectively
		usedRows.add(row)
		usedCols.add(col)

	# compute both the row and column index we have NOT yet
	# examined
	unusedRows = set(range(0, D.shape[0])).difference(usedRows)
	unusedCols = set(range(0, D.shape[1])).difference(usedCols)

	# loop over the unused row indexes
	for row in unusedRows:
		# grab the object ID for the corresponding row
		# index and increment the disappeared counter
		objectID = objectIDs[row]
		disappeared[objectID] += 1
		# check to see if the number of consecutive
		# frames the object has been marked "disappeared"
		# for warrants deregistering the object
		if disappeared[objectID] > MAX_DISAPPEARED:
			deregister(objectID)

	for col in unusedCols:
		register(curCentroids[col])


# takes in chromatophore contours, calculates area for each chromatophore
# adds area to appropriate area time series
# (the list of areas under the chrom w/ the correct ID num)
# returns updated chromAreas dict
def calcCurAreas(centroidsToContours, curFrameIndex):
	global chromAreas
	global matchedCentroids
	global disappeared
	global PIXELS_PER_MM

	# iterate through each centroid that's paired with an ID number
	# and add the area of each centroid's contour into chromAreas for this frame index
	for ID_num, centroid in matchedCentroids.items():
		# if the centroid is in the current frame, add its area to chromAreas
		# but if the centroid disappeared
		# then don't add a new area entry into chromAreas, just skip adding
		# the current frame as a data point for that chromatophore
		if disappeared[ID_num] < 1:
			curContour = centroidsToContours[centroid]  # get corresponding contour
			contourArea = cv2.contourArea(curContour)  # calculate area of corresponding contour

			# convert area from pixels^2 to mm^2
			contourArea = contourArea/(PIXELS_PER_MM**2)

			# add entry of area for this chrom's area in this frame
			chromAreas[ID_num][curFrameIndex] = contourArea


# draw ID numbers, contours or bounding boxes, and centroids for each object in the frame
# returns frame with everything drawn in it
def drawIDNums(shapes, frame, ROI):
	global matchedCentroids

	# draw rotated rectangles
	if args["ID_draw_shape"] == "rotated rectangles":
		# draw rotated rectangle bounding boxes
		for box in shapes:
			frame = drawRotatedRect(box, frame)
	# draw contours
	else:
		# draw contours on ID frame
		cv2.drawContours(image=frame, contours=shapes, contourIdx=-1, color=(0, 0, 255),
						 thickness=1)

	# draw centroids and ID nums
	for ID, centroid in matchedCentroids.items():
		# draw both the ID of the object and the centroid of the
		# object on the output frame
		text = str(ID)
		# offset numbers when a chromatophore is too close to the edges
		if centroid[0] > (ROI[2] - 17):  # too close to right
			text_x = int(centroid[0]) - 20
		elif centroid[0] < 7:  # too close to left
			text_x = int(centroid[0]) + 7
		else:
			text_x = int(centroid[0]) - 5
		if centroid[1] < 7:  # too close to top
			text_y = int(centroid[1]) + 5
		else:
			text_y = int(centroid[1]) - 5

		cv2.putText(frame, text, (text_x, text_y),
						cv2.FONT_HERSHEY_DUPLEX, 0.7, (255, 0, 0), 1)
		cv2.circle(frame, (int(centroid[0]), int(centroid[1])), 2, (0, 255, 0), -1)

	return frame


# takes in a 1D list of areas over time for a specific chromatphore, returns the standard deviation
def getStdDev(areas):
	# accounting for "divide by 0" error
	# if a "chromatophore" is so small that that the computed mm^2 is almost 0
	if sum(areas) == 0:
		return 0

	mean = sum(areas)/len(areas)
	variance = sum((x - mean) ** 2 for x in areas) / sum(areas)
	stdDev = math.sqrt(variance)
	return stdDev


# modified FROM METADATA.PY
# fill in gaps in data (essentially drawing a straight line between points across the gaps)
# area: list with area data (and None where no real data)
# just_nums: list of areas (when recorded, so not including the None entries)
# just_nums_frames: list of corresponding frames to the areas in just_nums
# returns list of same size, but with data interpolated
def fill_in_data_gaps(data, just_nums, just_nums_frames):
    filled_data = data.copy()

    # iterate through all the actual numerical data
    for i in range(len(just_nums_frames)):
        # if this is the first one in the just_nums_frames list
        # and there is a gap that at the start of the data, hold the first real num constant
        if i == 0 and just_nums_frames[i] != 0:
            for j in range(just_nums_frames[i]):
                filled_data[j] = just_nums[i]

        # if this is the last one in the just_num_frames list
        if i == len(just_nums_frames) - 1:
            # if there is a gap that lasts until the end of the vid, hold the last real num constant
            if just_nums_frames[i] != len(data) - 1:
                for j in range(just_nums_frames[i], len(data)):
                    filled_data[j] = just_nums[i]
            # if last frame has actual data in it, no need to do anything (no gaps to fill)
        # otherwise, fill gaps by connecting actual data points linearly
        else:
            cur_existing_f = just_nums_frames[i]  # existing frame num
            next_existing_f = just_nums_frames[i + 1]  # existing frame num

            # if there is a gap that starts at the cur existing frame
            if cur_existing_f + 1 != next_existing_f:
                # calculate number of blank frames (between the two existing frames)
                num_blank_frames = next_existing_f - cur_existing_f - 1
                # create filler datapoints (note: endpoints = actual data)
                filler_points = np.linspace(data[cur_existing_f], data[next_existing_f],
											num=num_blank_frames + 2)
                # fill in missing data in data array
                for j in range(cur_existing_f, next_existing_f + 1):
                    filled_data[j] = filler_points.item(j - cur_existing_f)

    return filled_data


# NOT CURRENTLY WORKING -- interpolation with numpy interpolate
# interpolate a chrom's area data
# areas: dict of frame : area
# last_frame: last frame (number) of entire video
def interpolateAreas(areas, last_frame):
	# get frame nums that don't have area data
	print(areas)
	missing_frames = [frame for frame in range(last_frame + 1) if frame not in areas.keys()]
	print("missing", len(missing_frames), "frames")

	# if no missing frames, nothing to interpolate. Return original
	if not missing_frames:
		return areas

	# if appears after start of vid, we'll hold the area constant until 1st real pt
	# do this by pretending that the 1st frame has the same area as the first actual area
	if missing_frames[0] == 0:
		first_actual_fr = min(areas.keys())
		for f in range(first_actual_fr):
			areas[f] = areas[first_actual_fr]  # set to most recently recorded area
			missing_frames.pop(f)  # remove frame from missing_frames since we pretend to see it

	# if disappears for rest of vid, we'll hold the area constant until end in interpolation
	# do this by pretending that the last frame has the same area as the most recently recorded one
	if missing_frames[-1] == last_frame:
		last_actual_fr = max(areas.keys())
		for f in range(last_frame, last_actual_fr, -1):
			areas[f] = areas[last_actual_fr]  # set to most recently recorded area
			missing_frames.remove(f)  # remove frame from missing_frames since we pretend to see it
		# areas[num_frames - 1] = areas[max(areas.keys())]  # set to most recently recorded area
		# missing_frames.pop(-1)  # remove last frame from missing_frames since we pretend to see it

	# if no missing frames, nothing to interpolate. Return original
	if not missing_frames:
		return areas

	print("now missing", len(missing_frames), "frames")
	print(missing_frames)
	print()
	print(areas.keys())
	print()
	print(areas.values())

	# evaluate the missing frames
	interpolated_area_points = np.interp(missing_frames, areas.keys(), areas.values())

	# add interpolated points to area time series dict
	interpolated = areas.copy()
	for i in range(len(missing_frames)):
		interpolated[missing_frames[i]] = interpolated_area_points[i]

	print(interpolated.keys(), "interpolated keys")
	print()
	print(interpolated.values(), "interpolated values")
	return interpolated


# filter out irrelevant data entries (ex: chrom that was only identified in
# a few frames)
# last_frame = index of last frame of the video
def cleanUpData(last_frame):
	global chromAreas

	# make a copy to iterate through to avoid "mutated during iteration" error
	origChromAreas = chromAreas.copy()

	for id_num, areas in origChromAreas.items():
		# if a certain chromatophore has less than 25 area data entries
		# (~1 sec since usually about 25fps), delete it from the dataset
		if len(areas) < 25:
			del chromAreas[id_num]
			del initialCentroids[id_num]
			continue

		# option 1
		# convert area time series dict to list
		areas_list = [areas[frame] if frame in areas.keys() else None for frame in range(last_frame + 1)]
		# interpolate chrom area
		interpolated_areas = fill_in_data_gaps(areas_list, list(areas.values()), list(areas.keys()))

		# option 2: numpy version -- not working -- ValueError: object of too small depth for desired array
		#interpolated_areas = interpolateAreas(areas, num_frames).values()

		# if a chromatophore doesn't change much over the course of the video, delete it
		#if getStdDev(areas.values()) < 0.05:  # option 3: old (didn't interpolate datapoints before getting std dev)
		if getStdDev(interpolated_areas) < 0.05:
			del chromAreas[id_num]
			del initialCentroids[id_num]


# formats final dataset into a spreadsheet and saves to the same folder as the video
# takes in number of frames in the video and the ROI coordinates/dimensions
# Important note: the xlwt library currently only supports up to 256 columns,
# so the amount of chromatophores in the selected ROI must be < 256
def formatData(numFrames, ROI, cleaned):
	global chromAreas

	# Workbook is created
	wb = Workbook()
	areasSheet = wb.create_sheet('Areas')
	areasSheet.title = "Areas"
	centroidsSheet = wb.create_sheet('Centroids')
	centroidsSheet.title = "Centroids"
	del wb['Sheet']  # delete default sheet

	# Put labels on the sheets
	areasSheet.cell(row=1, column=1).value = "frame"
	centroidsSheet.cell(row=1, column=1).value = "Centroid (pixels, relative to full image)"
	centroidsSheet.cell(row=2, column=1).value = "x"
	centroidsSheet.cell(row=3, column=1).value = "y"

	roiString = "_" + str(ROI[0]) + "_" + str(ROI[1]) + "_" + str(ROI[2]) + "_" + str(ROI[3])

	# add some useful sheet/vid-specific info
	centroidsSheet.cell(row=5, column=1).value = "Video:"
	centroidsSheet.cell(row=5, column=2).value = curVidPath
	centroidsSheet.cell(row=6, column=1).value = "ROI:"
	centroidsSheet.cell(row=6, column=2).value = roiString[1:]
	centroidsSheet.cell(row=7, column=1).value = "Pixels per mm:"
	centroidsSheet.cell(row=7, column=2).value = PIXELS_PER_MM

	# add column in column 0 that says each frame number
	for frame_num in range(numFrames + 1):
		areasSheet.cell(row=frame_num + 2, column=1).value = frame_num

	curCol = 2

	# iterate through each identified chromatophore
	for id_num, areas in chromAreas.items():
		# add ID num to both sheets on row 0
		areasSheet.cell(row=1, column=curCol).value = "C" + str(id_num)
		centroidsSheet.cell(row=1, column=curCol).value = "C" + str(id_num)

		# add centroid to the centroids sheet
		(x, y) = initialCentroids[id_num]
		centroidsSheet.cell(row=2, column=curCol).value = x
		centroidsSheet.cell(row=3, column=curCol).value = y

		# add area time series to areas sheet
		for frame_num, area in areas.items():
			areasSheet.cell(row=frame_num + 2, column=curCol).value = area

		curCol += 1

	if cleaned:
		directory = curVidPath.replace(vidName_ext, vidName_no_ext) + roiString + ".xlsx"
		wb.save(directory)
	else:
		directory = curVidPath.replace(vidName_ext, vidName_no_ext) + roiString + "_uncleaned" + ".xlsx"

	wb.save(directory)
	print("Saved xlsx data file to " + directory)


# Plots chrom areas over time (each chromatophore is one line) then shows the graph
# Commented out: saving the graph to the computer
def plotChromAreas():
	global chromAreas

	for id_num, areas in chromAreas.items():
		x = areas.keys()
		y = areas.values()

		plt.plot(x, y, label="C" + str(id_num))

	# label the graph
	plt.xlabel('Frame')  # naming the x-axis
	plt.ylabel('Area (in mm)')  # naming the y-axis
	plt.title('Areas of individual chromatophores for ' + curVidPath)  # title the graph
	plt.legend()  # show a legend on the plot

	# show the graph
	plt.show()

	# save the graph
	# curVid = curVidPath.replace("/", "_")
	# curVid = curVid.replace(".", "")
	# plt.savefig('areas_' + curVid + '.png')


# shows the current frame with chromatophore IDs, adjusting
# for mac and windows systems (running on windows caused an error with displaying the frame)
def showIDFrame(curFrameIndex, ID_frame):
	# resize frame when displaying
	display_frame = imutils.resize(ID_frame.copy(), width=(int(DISPLAY_SCALE*ROI_width)))
	cv2.imshow("ID frame", display_frame)  # display frame

	# uncomment if want frame nums on each window title (fine on mac but doesn't display properly on
	# windows)
	#cv2.imshow("ID frame for " + str(curFrameIndex), frame)
	# move window to same spot in screen as the previous windows
	#cv2.moveWindow("ID frame for " + str(curFrameIndex), 10, 10)

	# if the -s flag is raised, wait until the user presses a key to move onto the next frame
	if args["stop"] == "on":
		cv2.waitKey(0)

	cv2.waitKey(1)

	#if curFrameIndex > 0:
		# destroy previous frame's window to avoid program slowing down
		# uncomment this section if want frame number
		# displayed in window title, and switch out the first imshow line for the other imshow line
		# (currently commented out) in this showIDFrame function
		#cv2.destroyWindow("ID frame for " + str(curFrameIndex - 1))


# show images from each step of the process
def showAllImages(curFrameIndex, original, gray, threshold, contours, ID_labeled):
	cv2.imshow("Original" + str(curFrameIndex), original)
	cv2.imshow("Grayscale + Gaussian Blur" + str(curFrameIndex), gray)
	cv2.imshow("Threshold" + str(curFrameIndex), threshold)
	cv2.imshow("Contours" + str(curFrameIndex), contours)
	cv2.imshow("IDs + centroids + bounding boxes " + str(curFrameIndex), ID_labeled)

	cv2.waitKey(1)

	cv2.destroyWindow("Original" + str(curFrameIndex))
	cv2.destroyWindow("Grayscale + Gaussian Blur" + str(curFrameIndex))
	cv2.destroyWindow("Threshold" + str(curFrameIndex))
	cv2.destroyWindow("Contours" + str(curFrameIndex))
	cv2.destroyWindow("IDs + centroids + bounding boxes " + str(curFrameIndex))


# save image to frame_cap folder
def saveToFrameCap(ROI, curFrameIndex, img):
	roiString = "_" + str(ROI[0]) + "_" + str(ROI[1]) + "_" + str(ROI[2]) + "_" + str(ROI[3])
	frameCapDir = curVidPath.replace(vidName_ext, '') + "frame_cap_" + vidName_no_ext + roiString

	# checking if the frame_cap folder exists yet
	if not os.path.isdir(frameCapDir):
		# if the frame_cap directory is not present then create it
		os.makedirs(frameCapDir)

	# save in frame_cap folder
	cv2.imwrite(frameCapDir + "/process_" + str(curFrameIndex) + ".png", img)

	return frameCapDir


# saves one image that shows each step of the process for the current frame
# saves to the frame_cap folder
# to use, call at the end of the while True loop in processData()
def saveProcessImage(curFrameIndex, original, gray, threshold, contours, ID_labeled, ROI):
	# convert 2 channel images (grayscale) into 3 channel images (RGB)
	gray = cv2.cvtColor(gray, cv2.COLOR_GRAY2RGB)
	threshold = cv2.cvtColor(threshold, cv2.COLOR_GRAY2RGB)

	# stack images vertically
	processImg = np.concatenate((original, gray), axis=0)
	processImg = np.concatenate((processImg, threshold), axis=0)
	processImg = np.concatenate((processImg, contours), axis=0)
	processImg = np.concatenate((processImg, ID_labeled), axis=0)

	frameCapDir = saveToFrameCap(ROI, curFrameIndex, processImg)

	return frameCapDir


# had to make these into global variables to account for windows compatibility
ROI_x, ROI_y, ROI_width, ROI_height = 0, 0, ORIGINAL_WIDTH, ORIGINAL_HEIGHT


# save context image to show where the ROI is in the full frame
def saveRoiContextImg(frame, ROI):
	# draw rectangle for ROI
	rect_top_left = (ROI_x, ROI_y)
	rect_bottom_right = (ROI_x + ROI_width, ROI_y + ROI_height)
	cv2.rectangle(frame, rect_top_left, rect_bottom_right, (0, 0, 255), 2)

	# draw centroids and ID nums of filtered chroms
	for ID, centroid in initialCentroids.items():
		# draw both the ID of the object and the centroid of the
		# object on the output frame
		text = str(ID)

		x, y = int(centroid[0]), int(centroid[1])

		cv2.putText(frame, text, (x + 5, y - 5),
					cv2.FONT_HERSHEY_DUPLEX, 0.7, (255, 0, 0), 1)
		cv2.circle(frame, (int(x), int(y)), 2, (0, 255, 0), -1)

	roiString = "_" + str(ROI[0]) + "_" + str(ROI[1]) + "_" + str(ROI[2]) + "_" + str(ROI[3])
	directory = curVidPath.replace(vidName_ext, vidName_no_ext) + roiString + "_ROI_context.png"

	# save in frame_cap folder
	cv2.imwrite(directory, frame)
	print("Saved ROI context image to " + directory)


# summary function of entire program
# takes in filepath to raw video, goes through entire identification process,
# and saves the time series of each chromatophores' change in area as
# .xls files – one for cleaned data and one for uncleaned data
def processData(vidPath):
	global matchedCentroids
	global chromAreas
	global ROI_x
	global ROI_y
	global ROI_width
	global ROI_height

	# open the video
	vid = cv2.VideoCapture(vidPath)
	curFrameIndex = 0
	full_frame = None

	# run computer vision process for each frame in the video
	while True:
		# read current frame from the video
		ok, frame = vid.read()

		# stop loop if error reading frame or end of vid reached
		if not ok:
			break

		# select region of interest
		# note: x and y coordinates are of top left corner of ROI
		if curFrameIndex == 0:
			ROI_x, ROI_y, ROI_width, ROI_height = getROI(frame)
			ROI = [ROI_x, ROI_y, ROI_width, ROI_height]

			full_frame = frame

		# crop the frame to only include the ROI
		frame = frame[ROI_y:ROI_y + ROI_height, ROI_x: ROI_x + ROI_width]

		# mask out the chromatophores and display the masked frame
		gray_frame, thresh_frame = maskChromatophores(frame)

		# find contours from threshold
		contour_frame, contours = contour(thresh_frame, frame)

		# sort contours from left to right
		sortedContours, sortedBoundingBoxes = sortLR(contours)

		# put current centroids of each bounding box (unpaired with ID num) in a list
		curCentroids, centroidsToContours = getCentroids(sortedBoundingBoxes, sortedContours)

		# if first frame, create dict of chromatophores and begin object tracking
		if curFrameIndex == 0:
			createMatchedCentroids(curCentroids)  # initialize matchedCentroids dict
		# otherwise match each current centroid with an ID num
		else:
			matchCentroids(curCentroids)

		# draw matched ID numbers and centroids onto original frame
		if args["ID_draw_shape"] == "rotated rectangles":
			ID_frame = drawIDNums(sortedBoundingBoxes, frame.copy(), ROI)
		else:
			ID_frame = drawIDNums(contours, frame.copy(), ROI)

		if args["display_vid"] == "on":
			# show images
			#showAllImages(curFrameIndex, frame, gray_frame, thresh_frame, contour_frame, ID_frame)
			showIDFrame(curFrameIndex, ID_frame)

		# save generated image(s) to the frame_cap folder
		if args["frame_cap"] == "on, save entire process":
			frameCapDir = saveProcessImage(curFrameIndex, frame, gray_frame, thresh_frame,
										   contour_frame, ID_frame, ROI)
		elif args["frame_cap"] == "on, save ID frame":
			frameCapDir = saveToFrameCap(ROI, curFrameIndex, ID_frame)

		# add areas of chromatophores from current frame to chromAreas
		calcCurAreas(centroidsToContours, curFrameIndex)

		curFrameIndex += 1  # update frame index

	last_frame = curFrameIndex - 1

	if args["frame_cap"] != "off":
		roiString = "_" + str(ROI_x) + "_" + str(ROI_y) + "_" + str(ROI_width) + "_" + str(ROI_height)
		frameCapDir = curVidPath.replace(vidName_ext, '') + "frame_cap_" + vidName_no_ext + roiString
		print("Saved the frame captures in " + frameCapDir)

	if args["watch_only"] == "off":
		# save uncleaned data in a .xlsx file
		formatData(last_frame, ROI, cleaned=False)

		# filter out irrelevant data entries
		cleanUpData(last_frame)

		# format cleaned dataset and save as a .xls file in video's original directory
		formatData(curFrameIndex - 1, ROI, cleaned=True)

		# save ROI context image
		if args['save_ROI_context_img'] == "on":
			saveRoiContextImg(full_frame.copy(), ROI)

		# plot chromatophore areas
		#plotChromAreas()

		# close graph display
		#plt.close()

	# close any open windows
	cv2.destroyAllWindows()


def main():
	parse_args()

	# use argument vals if any given, otherwise use default nums
	setGlobalNums()

	# run the program's summary function
	processData(curVidPath)

	print("Program complete. Next, feed the generated .xlsx file into metadata.py")


# Using the special variable
# __name__
if __name__=="__main__":
	main()